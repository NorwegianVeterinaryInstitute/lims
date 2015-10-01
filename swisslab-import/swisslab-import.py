import sys
import itertools
from openpyxl import load_workbook
import StringIO
import requests
from genologics.lims import *
from genologics import config

# SwissLab data import script

# Use:
# python swisslab-import.py PROCESS-ID SWL-FILE
# Arguments:
#  PROCESS-ID: LIMS-ID of a process on which this script is run.
#  SWL-FILE:   LIMS-ID of Artifact object for Excel file

def get_swl_data(filename):
    """Parsing of sample information table in Excel format.
    
    Returns a dict indexed by sample name, where each value is a list 
    of (udfname, udfvalue) pairs.
    { sample name => [ (udfname, udfvalue), ... ] }
    """
    try:
        wb = load_workbook(filename, data_only=True)
    except IOError:
        print "Cannot read the SwissLab file, make sure it is in Excel format"
        sys.exit(1)

    ws = wb['Samples']
    udfnames = []

    assert ws['A1'].value == "Sample/Name"
    for i in itertools.count(2):
        h = ws.cell(column=i, row=1).value
        if h:
            udfnames.append(h)
        else:
            break

    done = False
    data = {}
    for row in itertools.count(2):
        sample_name = ws.cell(column=1, row=row).value
        if sample_name:
            sample_values = []
            for i, udfname in enumerate(udfnames):
                v = ws.cell(column=i+2, row=row).value
                sample_values.append((udfname, v))
            data[sample_name] = sample_values
        else:
            break

    return data
            

def main(process_id, swl_file_id):
    lims = Lims(config.BASEURI, config.USERNAME, config.PASSWORD)

    swl_file_art_obj = Artifact(lims, id=swl_file_id)
    if len(swl_file_art_obj.files) == 1:
        swl_io_obj = StringIO.StringIO(swl_file_art_obj.files[0].download())
    else:
        print "Could not access the SwissLab file, check that it has been uploaded"
        sys.exit(1)

    swisslab_data = get_swl_data(swl_io_obj)

    process = Process(lims, id=process_id)
    inputs = process.all_inputs(unique=True, resolve=True)
    assert all(len(input.samples) == 1 for input in inputs)
    samples = lims.get_batch(input.samples[0] for input in inputs)

    for sample in samples:
        try:
            fields = swisslab_data[sample.name]
        except KeyError:
            print "Failed trying to look up sample", sample.name, "in the provided table"
            sys.exit(1)

        for udfname, udfval in fields:
            print "on", sample.name, "setting", udfname, "to", udfval
            sample.udf[udfname] = str(udfval)

    lims.put_batch(samples)
    print "Imported data for", len(samples), "samples"


main(*sys.argv[1:])

